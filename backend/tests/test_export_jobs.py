import asyncio
import hashlib
import json
import time
import zipfile
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.models.audit_log import AuditAction, AuditLogEntry
from app.models.export_job import ExportJob, ExportJobStatus
from app.models.export_settings import ExportCleanupMode
from app.models.raw_event import RawEvent
from app.schemas.common import RangeParam
from app.services import export_job_service, export_settings_service


def _make_event(**overrides) -> RawEvent:
    defaults = dict(
        timestamp=datetime.now(UTC),
        duration_ms=1,
        client_ip="10.0.0.1",
        action="TCP_MISS",
        status_code=200,
        bytes=100,
        method="GET",
        url="http://example.com/",
        domain="example.com",
        user="alice",
        hierarchy="HIER_DIRECT",
        peer=None,
        content_type="text/html",
        blocked=False,
    )
    defaults.update(overrides)
    return RawEvent(**defaults)


async def test_run_job_writes_csv_and_marks_done(db_session: AsyncSession, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add_all([_make_event(client_ip=f"10.0.0.{i}") for i in range(5)])
    await db_session.commit()

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    assert job.status == ExportJobStatus.PENDING

    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.DONE
    assert refreshed.row_count == 5
    assert refreshed.file_path is not None
    assert refreshed.file_path.endswith(".zip")  # raw .csv is zipped and removed, see run_job
    with zipfile.ZipFile(refreshed.file_path) as zf:
        names = zf.namelist()
        assert len(names) == 1
        assert names[0] == export_job_service.inner_filename(refreshed)
        csv_text = zf.read(names[0]).decode()
    assert csv_text.count("\n") == 6  # header + 5 rows
    assert refreshed.file_size_bytes == Path(refreshed.file_path).stat().st_size
    assert refreshed.file_size_bytes > 0


async def test_create_job_records_export_created_audit_entry(db_session: AsyncSession):
    # Exporting raw traffic data is security-relevant the same way
    # user-management changes are (see create_job's comment) -- an admin
    # needs to be able to answer "who exported what, and when."
    since = RangeParam.ONE_HOUR.since()
    until = datetime.now(UTC)
    await export_job_service.create_job(db_session, since, until, "csv", True, "default", "actor-123")

    entry = (
        await db_session.execute(
            select(AuditLogEntry).where(AuditLogEntry.action == AuditAction.EXPORT_CREATED)
        )
    ).scalar_one()
    assert entry.actor_user_id == "actor-123"
    assert entry.target_user_id is None  # not about another account
    assert "format=csv" in entry.detail
    assert "blocked_only=True" in entry.detail
    assert "branch=default" in entry.detail


async def test_run_job_writes_valid_json(db_session: AsyncSession, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add_all([_make_event(client_ip=f"10.0.0.{i}") for i in range(3)])
    await db_session.commit()

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "json", False, None, "test-admin"
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.DONE
    with zipfile.ZipFile(refreshed.file_path) as zf:
        parsed = json.loads(zf.read(export_job_service.inner_filename(refreshed)))
    assert len(parsed) == 3


async def test_run_job_commits_progress_while_running(db_engine, tmp_path: Path, monkeypatch):
    # A wide range at real traffic volumes can take minutes; before this,
    # row_count stayed null for the entire RUNNING phase and only got set
    # once at the very end, so a client polling GET /export/jobs/{id} had
    # nothing to show for however long that took. Uses its own session
    # factory (rather than the single shared db_session other tests reuse)
    # specifically so the mid-run read below goes through a genuinely
    # separate session -- proving the progress update was actually
    # committed to the database, not just held in run_job's own session.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    session_factory = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", session_factory)

    async with session_factory() as setup_session:
        job = await export_job_service.create_job(
            setup_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
        )
    job_id = job.id

    # Forces every elapsed-time check in run_job's streaming loop to see
    # >= 1s having passed, so the (normally once-a-second) throttled commit
    # fires after every batch instead of only once at the very end.
    fake_clock = {"t": 0.0}

    def fake_monotonic() -> float:
        fake_clock["t"] += 2.0
        return fake_clock["t"]

    monkeypatch.setattr(time, "monotonic", fake_monotonic)

    mid_run_row_count = None

    async def _fake_stream(*args, **kwargs):
        row_counter = args[-1]
        row_counter[0] = 42
        yield "id,domain\n"

        nonlocal mid_run_row_count
        async with session_factory() as reader:
            mid_run_row_count = (await reader.get(ExportJob, job_id)).row_count

        row_counter[0] = 100
        yield "more,data\n"

    monkeypatch.setattr(export_job_service, "stream_csv", _fake_stream)

    await export_job_service.run_job(job_id)

    assert mid_run_row_count == 42  # visible to another session before the job finished

    async with session_factory() as verify_session:
        refreshed = await verify_session.get(ExportJob, job_id)
    assert refreshed.status == ExportJobStatus.DONE
    assert refreshed.row_count == 100


async def test_run_job_honors_cancellation_requested_while_pending(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    # Simulates a cancel request landing before run_job's task even starts --
    # request_cancellation only ever touches the in-memory flag, never the
    # job row itself, so nothing else needs to happen first.
    export_job_service.request_cancellation(job.id)

    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.CANCELLED
    assert list(tmp_path.iterdir()) == []  # never even started writing


async def test_run_job_honors_cancellation_requested_mid_stream(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    # Unlike the PENDING case above, this must clean up whatever partial
    # file streaming had already flushed before the cancellation checkpoint
    # was reached -- the same failure mode test_run_job_deletes_partial_file
    # _on_mid_stream_failure covers for an actual error.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )

    async def _fake_stream(*args, **kwargs):
        row_counter = args[-1]
        row_counter[0] = 7
        yield "id,domain\n"
        # request_cancellation is normally called from the cancel route
        # while run_job is mid-flight elsewhere -- from inside the fake
        # stream is the simplest way to land it exactly between batches.
        export_job_service.request_cancellation(job.id)
        yield "more,data\n"

    monkeypatch.setattr(export_job_service, "stream_csv", _fake_stream)

    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.CANCELLED
    assert refreshed.row_count == 7  # whatever had been written before the flag was seen
    assert list(tmp_path.iterdir()) == []  # partial file cleaned up, same as a genuine failure


async def test_run_job_marks_failed_on_error(db_session: AsyncSession, monkeypatch):
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)
    # No _jobs_dir patch -- the real EXPORT_JOBS_DIR default ("./export_jobs")
    # is harmless to create, so force a different failure instead: an
    # unwritable directory path.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: Path("/nonexistent/definitely/not/here"))

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.FAILED
    assert refreshed.error_message


async def test_run_job_deletes_partial_file_on_mid_stream_failure(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    # A failure that happens *after* some chunks were already flushed to
    # disk (unlike test_run_job_marks_failed_on_error above, which fails
    # before ever opening the file) used to leave a truncated file behind
    # forever: job.file_path is only ever set on the success path, so
    # purge_old_jobs (which only unlinks via job.file_path) could never
    # find it.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    async def _broken_stream(*args, **kwargs):
        yield "id,domain\n"
        raise RuntimeError("simulated mid-stream failure")

    monkeypatch.setattr(export_job_service, "stream_csv", _broken_stream)

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.FAILED
    assert list(tmp_path.iterdir()) == []


async def test_reconcile_orphaned_jobs_fails_stale_pending_and_running_jobs(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    # Simulates what a crash/restart mid-export leaves behind: rows stuck in
    # PENDING/RUNNING with no asyncio task left anywhere to ever move them
    # forward (the one that would have belonged to the previous process),
    # plus the partial file the interrupted run_job had started writing.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    now = datetime.now(UTC)
    orphaned_running_file = tmp_path / "orphaned-running.csv"
    orphaned_running_file.write_text("id,domain\n1,example.com\n")
    db_session.add_all(
        [
            ExportJob(
                id="orphaned-pending",
                status=ExportJobStatus.PENDING,
                format="csv",
                since=now - timedelta(hours=1),
                until=now,
                blocked_only=False,
                branch=None,
                created_at=now,
            ),
            ExportJob(
                id="orphaned-running",
                status=ExportJobStatus.RUNNING,
                format="csv",
                since=now - timedelta(hours=1),
                until=now,
                blocked_only=False,
                branch=None,
                created_at=now,
            ),
            ExportJob(
                id="already-done",
                status=ExportJobStatus.DONE,
                format="csv",
                since=now - timedelta(hours=1),
                until=now,
                blocked_only=False,
                branch=None,
                file_path=str(tmp_path / "already-done.csv"),
                row_count=0,
                created_at=now,
                completed_at=now,
            ),
        ]
    )
    await db_session.commit()

    reconciled_count = await export_job_service.reconcile_orphaned_jobs()

    assert reconciled_count == 2
    pending = await db_session.get(ExportJob, "orphaned-pending")
    running = await db_session.get(ExportJob, "orphaned-running")
    done = await db_session.get(ExportJob, "already-done")
    assert pending.status == ExportJobStatus.FAILED
    assert running.status == ExportJobStatus.FAILED
    assert running.error_message
    assert done.status == ExportJobStatus.DONE  # untouched
    assert not orphaned_running_file.exists()


async def test_purge_old_jobs_deletes_expired_rows_and_files(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    old_file = tmp_path / "old.csv"
    old_file.write_text("id,domain\n")
    recent_file = tmp_path / "recent.csv"
    recent_file.write_text("id,domain\n")

    # Explicit, rather than relying on export_settings_service's default --
    # keeps this test's pass/fail independent of whatever that default
    # happens to be tuned to.
    await export_settings_service.update_settings(
        db_session,
        cleanup_mode=ExportCleanupMode.TIME_BASED,
        retention_hours=48,
        warn_undownloaded_after_hours=None,
        actor_user_id="actor-1",
    )

    now = datetime.now(UTC)
    db_session.add_all(
        [
            ExportJob(
                id="old-job",
                status=ExportJobStatus.DONE,
                format="csv",
                since=now - timedelta(days=1),
                until=now,
                blocked_only=False,
                branch=None,
                file_path=str(old_file),
                created_at=now - timedelta(hours=49),
            ),
            ExportJob(
                id="recent-job",
                status=ExportJobStatus.DONE,
                format="csv",
                since=now - timedelta(days=1),
                until=now,
                blocked_only=False,
                branch=None,
                file_path=str(recent_file),
                created_at=now - timedelta(hours=1),
            ),
        ]
    )
    await db_session.commit()

    deleted_count = await export_job_service.purge_old_jobs(db_session)
    await db_session.commit()

    assert deleted_count == 1
    assert await db_session.get(ExportJob, "old-job") is None
    assert await db_session.get(ExportJob, "recent-job") is not None
    assert not old_file.exists()
    assert recent_file.exists()


async def test_purge_old_jobs_is_noop_in_after_download_mode(db_session: AsyncSession, tmp_path: Path):
    """AFTER_DOWNLOAD mode deletes a job's file individually, right after
    each download (see delete_file_if_after_download_mode) -- purge_old_jobs
    must never also delete an old, still-undownloaded job by age in this
    mode, since that would silently destroy data the admin was relying on
    this mode specifically to keep around until they actually grab it."""
    old_file = tmp_path / "old.csv"
    old_file.write_text("id,domain\n")

    now = datetime.now(UTC)
    db_session.add(
        ExportJob(
            id="old-undownloaded-job",
            status=ExportJobStatus.DONE,
            format="csv",
            since=now - timedelta(days=1),
            until=now,
            blocked_only=False,
            branch=None,
            file_path=str(old_file),
            created_at=now - timedelta(hours=200),
        )
    )
    await export_settings_service.update_settings(
        db_session,
        ExportCleanupMode.AFTER_DOWNLOAD,
        retention_hours=48,
        warn_undownloaded_after_hours=None,
        actor_user_id="actor-1",
    )

    deleted_count = await export_job_service.purge_old_jobs(db_session)

    assert deleted_count == 0
    assert await db_session.get(ExportJob, "old-undownloaded-job") is not None
    assert old_file.exists()


async def test_record_download_sets_downloaded_at_only_on_first_call(db_session: AsyncSession):
    now = datetime.now(UTC)
    job = ExportJob(
        id="dl-job",
        status=ExportJobStatus.DONE,
        format="csv",
        since=now - timedelta(hours=1),
        until=now,
        blocked_only=False,
        branch=None,
        file_path=None,
        created_at=now,
    )
    db_session.add(job)
    await db_session.commit()

    assert job.downloaded_at is None

    await export_job_service.record_download(db_session, job, "admin-user-id")
    first_downloaded_at = job.downloaded_at
    assert first_downloaded_at is not None

    await export_job_service.record_download(db_session, job, "admin-user-id")
    assert job.downloaded_at == first_downloaded_at


async def test_delete_file_if_after_download_mode_deletes_only_in_that_mode(
    db_engine, tmp_path: Path, monkeypatch
):
    session_factory = async_sessionmaker(db_engine, expire_on_commit=False, class_=AsyncSession)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", session_factory)

    result_file = tmp_path / "result.zip"
    result_file.write_text("fake zip contents")

    now = datetime.now(UTC)
    async with session_factory() as session:
        session.add(
            ExportJob(
                id="after-dl-job",
                status=ExportJobStatus.DONE,
                format="csv",
                since=now - timedelta(hours=1),
                until=now,
                blocked_only=False,
                branch=None,
                file_path=str(result_file),
                created_at=now,
            )
        )
        await export_settings_service.update_settings(
            session,
            ExportCleanupMode.TIME_BASED,
            retention_hours=48,
            warn_undownloaded_after_hours=None,
            actor_user_id="actor-1",
        )

    # TIME_BASED (the default/current mode) -- must not touch the file.
    await export_job_service.delete_file_if_after_download_mode("after-dl-job")
    assert result_file.exists()

    async with session_factory() as session:
        await export_settings_service.update_settings(
            session,
            ExportCleanupMode.AFTER_DOWNLOAD,
            retention_hours=48,
            warn_undownloaded_after_hours=None,
            actor_user_id="actor-1",
        )

    await export_job_service.delete_file_if_after_download_mode("after-dl-job")
    assert not result_file.exists()

    async with session_factory() as session:
        job = await session.get(ExportJob, "after-dl-job")
        assert job.file_path is None


async def test_create_export_job_route_returns_pending(
    app_client: AsyncClient, admin_token, auth_headers, tmp_path: Path, monkeypatch
):
    # Keeps the real background task (still fired for real here) from
    # writing into the repo's working directory during the test.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)

    response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    assert response.status_code == 201
    body = response.json()
    assert body["status"] in ("pending", "running", "done")  # the real background task may race ahead
    assert body["format"] == "csv"


async def test_download_export_job_route_records_audit_entry(
    app_client: AsyncClient, admin_token, auth_headers, tmp_path: Path, monkeypatch, db_session: AsyncSession
):
    # EXPORT_CREATED (see test_create_job_records_export_created_audit_entry
    # above) only proves someone queued the export -- jobs are visible to
    # every admin, not just whoever created them, so this is the entry that
    # proves who actually pulled the finished file down.
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)

    create_response = await app_client.post(
        "/api/export/jobs?range=1h&format=csv", headers=auth_headers(admin_token)
    )
    job_id = create_response.json()["id"]

    for _ in range(20):
        status_response = await app_client.get(
            f"/api/export/jobs/{job_id}", headers=auth_headers(admin_token)
        )
        if status_response.json()["status"] == "done":
            break
        await asyncio.sleep(0.05)
    else:
        raise AssertionError("export job never reached done")

    download_response = await app_client.get(
        f"/api/export/jobs/{job_id}/download", headers=auth_headers(admin_token)
    )
    assert download_response.status_code == 200

    entry = (
        await db_session.execute(
            select(AuditLogEntry).where(AuditLogEntry.action == AuditAction.EXPORT_DOWNLOADED)
        )
    ).scalar_one()
    assert entry.actor_email == "admin@example.com"
    assert f"job_id={job_id}" in entry.detail


async def test_download_export_job_route_deletes_file_in_after_download_mode(
    app_client: AsyncClient, admin_token, auth_headers, tmp_path: Path, monkeypatch, db_session: AsyncSession
):
    """End-to-end through the real route (not calling the service function
    directly) -- proves the FastAPI BackgroundTask actually runs after the
    file is streamed, using the same DB the route itself writes through."""
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    await export_settings_service.update_settings(
        db_session,
        ExportCleanupMode.AFTER_DOWNLOAD,
        retention_hours=48,
        warn_undownloaded_after_hours=None,
        actor_user_id="actor-1",
    )

    create_response = await app_client.post(
        "/api/export/jobs?range=1h&format=csv", headers=auth_headers(admin_token)
    )
    job_id = create_response.json()["id"]

    for _ in range(20):
        status_response = await app_client.get(
            f"/api/export/jobs/{job_id}", headers=auth_headers(admin_token)
        )
        if status_response.json()["status"] == "done":
            break
        await asyncio.sleep(0.05)
    else:
        raise AssertionError("export job never reached done")

    download_response = await app_client.get(
        f"/api/export/jobs/{job_id}/download", headers=auth_headers(admin_token)
    )
    assert download_response.status_code == 200

    job = await db_session.get(ExportJob, job_id)
    await db_session.refresh(job)
    assert job.file_path is None
    assert job.downloaded_at is not None

    # A second attempt must fail cleanly (409, matching the "not ready
    # yet"/already-gone case), not crash trying to serve a deleted file.
    second_response = await app_client.get(
        f"/api/export/jobs/{job_id}/download", headers=auth_headers(admin_token)
    )
    assert second_response.status_code == 409


async def test_download_export_job_route_409_does_not_record_audit_entry(
    app_client: AsyncClient, admin_token, auth_headers, monkeypatch, db_session: AsyncSession
):
    # The 409-before-done case (see test_download_export_job_route_409_
    # before_done below) never actually serves the file -- nothing was
    # downloaded, so nothing should be audited.
    monkeypatch.setattr(export_job_service, "start", lambda job_id: None)

    create_response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    job_id = create_response.json()["id"]

    response = await app_client.get(f"/api/export/jobs/{job_id}/download", headers=auth_headers(admin_token))
    assert response.status_code == 409

    count = (
        (
            await db_session.execute(
                select(AuditLogEntry).where(AuditLogEntry.action == AuditAction.EXPORT_DOWNLOADED)
            )
        )
        .scalars()
        .all()
    )
    assert count == []


async def test_get_export_job_route_404_for_unknown_id(app_client: AsyncClient, admin_token, auth_headers):
    response = await app_client.get("/api/export/jobs/does-not-exist", headers=auth_headers(admin_token))
    assert response.status_code == 404


async def test_download_export_job_route_409_before_done(
    app_client: AsyncClient, admin_token, auth_headers, monkeypatch
):
    # Prevent the real background task from racing to completion during
    # the test, so the job is deterministically still PENDING/RUNNING.
    monkeypatch.setattr(export_job_service, "start", lambda job_id: None)

    create_response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    job_id = create_response.json()["id"]

    response = await app_client.get(f"/api/export/jobs/{job_id}/download", headers=auth_headers(admin_token))
    assert response.status_code == 409


async def test_cancel_export_job_route_marks_job_cancelled(
    app_client: AsyncClient, admin_token, auth_headers, monkeypatch
):
    # Prevent the real background task from racing to completion, same as
    # test_download_export_job_route_409_before_done above -- keeps the job
    # deterministically PENDING so the cancel endpoint's own effect (not a
    # race with run_job finishing on its own) is what's being tested.
    monkeypatch.setattr(export_job_service, "start", lambda job_id: None)

    create_response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    job_id = create_response.json()["id"]

    cancel_response = await app_client.post(
        f"/api/export/jobs/{job_id}/cancel", headers=auth_headers(admin_token)
    )
    assert cancel_response.status_code == 200
    assert cancel_response.json()["status"] == "pending"  # cancellation is async -- see the route's docstring

    # Actually running run_job now (start() was stubbed out above) proves
    # the cancel request was recorded and gets honored once the job's own
    # task next checks for it -- immediately, since it's still PENDING.
    await export_job_service.run_job(job_id)
    status_response = await app_client.get(f"/api/export/jobs/{job_id}", headers=auth_headers(admin_token))
    assert status_response.json()["status"] == "cancelled"


async def test_cancel_export_job_route_records_audit_entry(
    app_client: AsyncClient, admin_token, auth_headers, db_session: AsyncSession, monkeypatch
):
    monkeypatch.setattr(export_job_service, "start", lambda job_id: None)

    create_response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    job_id = create_response.json()["id"]

    cancel_response = await app_client.post(
        f"/api/export/jobs/{job_id}/cancel", headers=auth_headers(admin_token)
    )
    assert cancel_response.status_code == 200

    entry = (
        await db_session.execute(
            select(AuditLogEntry).where(AuditLogEntry.action == AuditAction.EXPORT_CANCELLED)
        )
    ).scalar_one()
    assert job_id in entry.detail


async def test_cancel_export_job_route_409_when_already_done(
    app_client: AsyncClient, admin_token, auth_headers
):
    create_response = await app_client.post(
        "/api/export/jobs?range=1h&format=csv", headers=auth_headers(admin_token)
    )
    job_id = create_response.json()["id"]
    # No data in range -- the real background task (still fired for real
    # here) reaches DONE almost immediately.
    for _ in range(20):
        status_response = await app_client.get(
            f"/api/export/jobs/{job_id}", headers=auth_headers(admin_token)
        )
        if status_response.json()["status"] == "done":
            break
        await asyncio.sleep(0.05)
    else:
        raise AssertionError("export job never reached done")

    cancel_response = await app_client.post(
        f"/api/export/jobs/{job_id}/cancel", headers=auth_headers(admin_token)
    )
    assert cancel_response.status_code == 409


async def test_cancel_export_job_route_404_for_unknown_id(app_client: AsyncClient, admin_token, auth_headers):
    response = await app_client.post(
        "/api/export/jobs/does-not-exist/cancel", headers=auth_headers(admin_token)
    )
    assert response.status_code == 404


async def test_list_export_jobs_route_requires_admin(app_client: AsyncClient, viewer_token, auth_headers):
    response = await app_client.get("/api/export/jobs", headers=auth_headers(viewer_token))
    assert response.status_code == 403


async def test_create_export_job_route_429_at_max_concurrent(
    app_client: AsyncClient, admin_token, auth_headers, monkeypatch
):
    # Nothing else caps how many exports can run at once, and each can be a
    # multi-hundred-MB file running for minutes -- without this, several
    # tabs/admins kicking off wide-range exports at the same time could fill
    # the disk. Keeps jobs stuck PENDING (like test_download_..._409_before_done
    # above) so hitting the default EXPORT_JOB_MAX_CONCURRENT (3) is deterministic.
    monkeypatch.setattr(export_job_service, "start", lambda job_id: None)

    for _ in range(3):
        response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
        assert response.status_code == 201

    response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    assert response.status_code == 429


async def test_create_job_stores_and_echoes_filters_and_columns(db_session: AsyncSession):
    job = await export_job_service.create_job(
        db_session,
        RangeParam.ONE_HOUR.since(),
        datetime.now(UTC),
        "csv",
        False,
        None,
        "test-admin",
        client_ip="10.0.0.5",
        domain="example.com",
        category="social_media",
        columns=["id", "domain"],
    )

    assert job.client_ip == "10.0.0.5"
    assert job.domain == "example.com"
    assert job.category == "social_media"
    assert job.columns == "id,domain"

    out = export_job_service.to_out(job)
    assert out.client_ip == "10.0.0.5"
    assert out.columns == ["id", "domain"]


async def test_create_job_rejects_unknown_column(db_session: AsyncSession):
    with pytest.raises(ValueError, match="not-a-real-column"):
        await export_job_service.create_job(
            db_session,
            RangeParam.ONE_HOUR.since(),
            datetime.now(UTC),
            "csv",
            False,
            None,
            "test-admin",
            columns=["not-a-real-column"],
        )


async def test_run_job_respects_client_ip_filter(db_session: AsyncSession, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add_all([_make_event(client_ip="10.0.0.1"), _make_event(client_ip="10.0.0.2")])
    await db_session.commit()

    job = await export_job_service.create_job(
        db_session,
        RangeParam.ONE_HOUR.since(),
        datetime.now(UTC),
        "csv",
        False,
        None,
        "test-admin",
        client_ip="10.0.0.2",
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.DONE
    assert refreshed.row_count == 1


async def test_run_job_writes_xlsx_and_marks_done(db_session: AsyncSession, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add_all([_make_event(client_ip=f"10.0.0.{i}") for i in range(4)])
    await db_session.commit()

    job = await export_job_service.create_job(
        db_session,
        RangeParam.ONE_HOUR.since(),
        datetime.now(UTC),
        "xlsx",
        False,
        None,
        "test-admin",
        columns=["client_ip", "domain"],
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.DONE
    assert refreshed.row_count == 4
    # No outer zip for xlsx (see result_filename's docstring).
    assert refreshed.file_path.endswith(".xlsx")
    assert refreshed.checksum_sha256 is not None
    assert len(refreshed.checksum_sha256) == 64  # sha256 hex digest length

    workbook = load_workbook(refreshed.file_path, read_only=True)
    rows = list(workbook["events"].iter_rows(values_only=True))
    assert rows[0] == ("client_ip", "domain")
    assert len(rows) == 5  # header + 4 rows


async def test_run_job_writes_xlsx_with_control_characters_in_url(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    """A url/user/etc. containing an ASCII control character (a malformed
    request, or someone probing with control bytes) used to fail the whole
    job -- openpyxl's IllegalCharacterError, raised the moment
    worksheet.append() hit the bad value, surfaced to the user as a bare
    "cannot be used in worksheets" error with no indication of which row.
    The bad byte is now stripped instead of the export failing."""
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add(_make_event(url="http://example.com/\x00bad\x1fpath", domain="example.com"))
    await db_session.commit()

    job = await export_job_service.create_job(
        db_session,
        RangeParam.ONE_HOUR.since(),
        datetime.now(UTC),
        "xlsx",
        False,
        None,
        "test-admin",
        columns=["url", "domain"],
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.status == ExportJobStatus.DONE
    assert refreshed.error_message is None

    workbook = load_workbook(refreshed.file_path, read_only=True)
    rows = list(workbook["events"].iter_rows(values_only=True))
    assert rows[1] == ("http://example.com/badpath", "example.com")


async def test_run_job_sets_checksum_for_csv(db_session: AsyncSession, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add(_make_event())
    await db_session.commit()

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    await export_job_service.run_job(job.id)

    refreshed = await db_session.get(ExportJob, job.id)
    assert refreshed.checksum_sha256 == hashlib.sha256(Path(refreshed.file_path).read_bytes()).hexdigest()


async def test_share_link_create_verify_and_revoke(db_session: AsyncSession, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    await export_job_service.run_job(job.id)
    job = await db_session.get(ExportJob, job.id)

    link = await export_job_service.create_share_link(db_session, job, "admin-1")
    assert link.token
    assert job.share_token_hash is not None

    # The right token verifies...
    verified = await export_job_service.verify_share_token(db_session, job.id, link.token)
    assert verified is not None
    assert verified.id == job.id

    # ...a wrong one doesn't.
    assert await export_job_service.verify_share_token(db_session, job.id, "wrong-token") is None

    # An expired one doesn't either.
    job.share_token_expires_at = datetime.now(UTC) - timedelta(seconds=1)
    await db_session.commit()
    assert await export_job_service.verify_share_token(db_session, job.id, link.token) is None

    # Revoking clears the link entirely, regardless of expiry.
    job.share_token_expires_at = datetime.now(UTC) + timedelta(hours=1)
    await db_session.commit()
    await export_job_service.revoke_share_link(db_session, job, "actor-1")
    assert job.share_token_hash is None
    assert await export_job_service.verify_share_token(db_session, job.id, link.token) is None

    entry = (
        await db_session.execute(
            select(AuditLogEntry).where(AuditLogEntry.action == AuditAction.EXPORT_SHARE_REVOKED)
        )
    ).scalar_one()
    assert entry.actor_user_id == "actor-1"
    assert job.id in entry.detail


async def test_create_share_link_records_export_shared_audit_entry(
    db_session: AsyncSession, tmp_path: Path, monkeypatch
):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, None, "test-admin"
    )
    await export_job_service.run_job(job.id)
    job = await db_session.get(ExportJob, job.id)

    await export_job_service.create_share_link(db_session, job, "admin-1")

    entry = (
        await db_session.execute(
            select(AuditLogEntry).where(AuditLogEntry.action == AuditAction.EXPORT_SHARED)
        )
    ).scalar_one()
    assert entry.actor_user_id == "admin-1"
    assert f"job_id={job.id}" in entry.detail


async def test_shared_download_route_works_with_valid_token(
    app_client: AsyncClient, admin_token, auth_headers, tmp_path: Path, monkeypatch
):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)

    create_response = await app_client.post(
        "/api/export/jobs?range=1h&format=csv", headers=auth_headers(admin_token)
    )
    job_id = create_response.json()["id"]
    for _ in range(20):
        status_response = await app_client.get(
            f"/api/export/jobs/{job_id}", headers=auth_headers(admin_token)
        )
        if status_response.json()["status"] == "done":
            break
        await asyncio.sleep(0.05)
    else:
        raise AssertionError("export job never reached done")

    share_response = await app_client.post(
        f"/api/export/jobs/{job_id}/share", headers=auth_headers(admin_token)
    )
    assert share_response.status_code == 200
    token = share_response.json()["token"]

    # No Authorization header at all -- the token alone is the credential.
    download_response = await app_client.get(
        f"/api/export/jobs/{job_id}/shared-download", params={"token": token}
    )
    assert download_response.status_code == 200


async def test_shared_download_route_rejects_invalid_token(
    app_client: AsyncClient, admin_token, auth_headers, tmp_path: Path, monkeypatch
):
    monkeypatch.setattr(export_job_service, "_jobs_dir", lambda: tmp_path)

    create_response = await app_client.post(
        "/api/export/jobs?range=1h&format=csv", headers=auth_headers(admin_token)
    )
    job_id = create_response.json()["id"]

    response = await app_client.get(
        f"/api/export/jobs/{job_id}/shared-download", params={"token": "definitely-not-valid"}
    )
    assert response.status_code == 404


async def test_share_export_job_route_404_before_ready(
    app_client: AsyncClient, admin_token, auth_headers, monkeypatch
):
    monkeypatch.setattr(export_job_service, "start", lambda job_id: None)

    create_response = await app_client.post("/api/export/jobs?format=csv", headers=auth_headers(admin_token))
    job_id = create_response.json()["id"]

    response = await app_client.post(f"/api/export/jobs/{job_id}/share", headers=auth_headers(admin_token))
    assert response.status_code == 409


# --- Branch scoping: a branch-scoped admin must not be able to see, poll,
# cancel, download, share, or revoke another branch's export job by ID --
# regression coverage for the IDOR export/jobs routes used to have zero
# branch enforcement for. ---


async def _create_branch_b_job(db_session: AsyncSession) -> str:
    job = await export_job_service.create_job(
        db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), "csv", False, "branch-b", "some-admin-id"
    )
    return job.id


async def test_branch_admin_list_export_jobs_excludes_other_branches(
    app_client: AsyncClient, branch_a_admin_token, auth_headers, db_session: AsyncSession
):
    await _create_branch_b_job(db_session)

    response = await app_client.get("/api/export/jobs", headers=auth_headers(branch_a_admin_token))
    assert response.status_code == 200
    assert response.json() == []


async def test_branch_admin_cannot_get_other_branch_job(
    app_client: AsyncClient, branch_a_admin_token, auth_headers, db_session: AsyncSession
):
    job_id = await _create_branch_b_job(db_session)

    response = await app_client.get(
        f"/api/export/jobs/{job_id}", headers=auth_headers(branch_a_admin_token)
    )
    assert response.status_code == 404


async def test_branch_admin_cannot_cancel_other_branch_job(
    app_client: AsyncClient, branch_a_admin_token, auth_headers, db_session: AsyncSession
):
    job_id = await _create_branch_b_job(db_session)

    response = await app_client.post(
        f"/api/export/jobs/{job_id}/cancel", headers=auth_headers(branch_a_admin_token)
    )
    assert response.status_code == 404


async def test_branch_admin_cannot_download_other_branch_job(
    app_client: AsyncClient, branch_a_admin_token, auth_headers, db_session: AsyncSession
):
    job_id = await _create_branch_b_job(db_session)

    response = await app_client.get(
        f"/api/export/jobs/{job_id}/download", headers=auth_headers(branch_a_admin_token)
    )
    assert response.status_code == 404


async def test_branch_admin_cannot_share_other_branch_job(
    app_client: AsyncClient, branch_a_admin_token, auth_headers, db_session: AsyncSession
):
    job_id = await _create_branch_b_job(db_session)

    response = await app_client.post(
        f"/api/export/jobs/{job_id}/share", headers=auth_headers(branch_a_admin_token)
    )
    assert response.status_code == 404


async def test_branch_admin_cannot_revoke_other_branch_job_share(
    app_client: AsyncClient, branch_a_admin_token, auth_headers, db_session: AsyncSession
):
    job_id = await _create_branch_b_job(db_session)

    response = await app_client.post(
        f"/api/export/jobs/{job_id}/share/revoke", headers=auth_headers(branch_a_admin_token)
    )
    assert response.status_code == 404
