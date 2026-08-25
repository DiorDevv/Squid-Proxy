"""CSV/JSON/XLSX export of raw events for a time range. Admin-only (see api/routes/export.py)."""

import csv
import io
import json
from collections.abc import AsyncIterator, Sequence
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import db as db_module
from app.models.domain_category import DomainCategoryLabel
from app.models.raw_event import RawEvent
from app.services.category_inference import effective_category
from app.services.domain_category_service import get_overrides_map
from app.services.event_query_service import build_event_conditions, distinct_domains

# Public (not `_`-prefixed): api/routes/export.py validates a requested
# `columns` list against this before it ever reaches a query, and
# schemas/export.py echoes it back for the frontend's column picker -- one
# canonical order/spelling every layer agrees on, rather than each
# re-declaring its own copy that could quietly drift from this one.
EXPORT_COLUMNS = [
    "id",
    "timestamp",
    "client_ip",
    "branch",
    "user",
    "method",
    "url",
    "domain",
    "action",
    "status_code",
    "bytes",
    "blocked",
]

EXPORT_ROW_LIMIT = 100_000


def resolve_columns(columns: list[str] | None) -> list[str]:
    """`columns=None` (the long-standing default) means every column, in
    EXPORT_COLUMNS' canonical order -- a caller-supplied list is used
    verbatim (including its order) but only after every entry is confirmed
    to be a real, known column. Never trust it straight from a query string:
    it ends up as `csv.DictWriter(fieldnames=...)`, and an unrecognized name
    there would fail with a confusing KeyError deep inside the streaming
    loop instead of a clean 400 at the API boundary (see api/routes/export.py)."""
    if columns is None:
        return list(EXPORT_COLUMNS)
    unknown = [c for c in columns if c not in EXPORT_COLUMNS]
    if unknown:
        raise ValueError(f"Unknown export column(s): {', '.join(unknown)}")
    if not columns:
        raise ValueError("At least one export column must be selected.")
    return columns


async def resolve_category_domains(
    session: AsyncSession,
    since: datetime,
    until: datetime,
    category: DomainCategoryLabel,
    branch: str | None = None,
) -> set[str]:
    """Turns a `category` filter into the concrete domain set
    build_event_conditions' `domain_in` actually filters on -- which
    category a domain falls under is Python business logic
    (category_inference.py), not something expressible as SQL, the same
    reason stats_service.get_top_domains resolves it in a separate pass
    rather than in its own query (see that function's docstring)."""
    overrides = await get_overrides_map(session)
    domains = await distinct_domains(session, since, until, branch)
    return {domain for domain in domains if effective_category(domain, overrides) == category}


# Columns that can carry attacker-influenced text (from Squid log lines) and
# so need CSV formula-injection escaping before being opened in a spreadsheet.
_FORMULA_RISK_COLUMNS = {"client_ip", "user", "url", "domain"}
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _escape_csv_formula(value: object) -> object:
    """Neutralize CSV/Excel formula injection (leading =, +, -, @, tab, CR).

    Excel/Sheets treat a cell as a formula if it starts with one of these
    characters, regardless of column type; prefixing with a single quote
    forces it to be read back as literal text.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return f"'{value}"
    return value


async def _fetch_rows(
    session: AsyncSession, since: datetime, until: datetime, blocked_only: bool, branch: str | None = None
) -> Sequence[RawEvent]:
    conditions = build_event_conditions(since, until, blocked_only=blocked_only, branch=branch)

    query = (
        select(RawEvent)
        .where(*conditions)
        .order_by(RawEvent.timestamp.desc())
        .limit(EXPORT_ROW_LIMIT)
    )
    return (await session.execute(query)).scalars().all()


def _escape_record(record: dict) -> dict:
    for column in _FORMULA_RISK_COLUMNS:
        if column in record:
            record[column] = _escape_csv_formula(record[column])
    return record


def _row_to_dict(row: RawEvent) -> dict:
    return {
        "id": row.id,
        "timestamp": row.timestamp.isoformat() if isinstance(row.timestamp, datetime) else row.timestamp,
        "client_ip": row.client_ip,
        "branch": row.branch,
        "user": row.user,
        "method": row.method,
        "url": row.url,
        "domain": row.domain,
        "action": row.action,
        "status_code": row.status_code,
        "bytes": row.bytes,
        "blocked": row.blocked,
    }


async def export_as_csv(
    session: AsyncSession, since: datetime, until: datetime, blocked_only: bool, branch: str | None = None
) -> str:
    rows = await _fetch_rows(session, since, until, blocked_only, branch)
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow(_escape_record(_row_to_dict(row)))
    return buffer.getvalue()


async def export_as_json(
    session: AsyncSession, since: datetime, until: datetime, blocked_only: bool, branch: str | None = None
) -> str:
    rows = await _fetch_rows(session, since, until, blocked_only, branch)
    return json.dumps([_row_to_dict(row) for row in rows])


# --- Streaming, uncapped variants ---
#
# The two functions above are for report_service.py's emailed CSV attachment
# (EXPORT_ROW_LIMIT and an in-memory string are both fine there -- nobody
# wants a multi-hundred-MB email attachment anyway). GET /api/export and
# scripts/archive_weekly_export.py need the opposite: the *complete* range
# even when that's millions of rows (a real deployment's raw_events table
# holds that much for even a single day), without ever holding the whole
# thing in memory at once. Keyset pagination (by id, not OFFSET) keeps every
# query O(batch size) regardless of how far into the range it is.

_STREAM_BATCH_SIZE = 5_000


async def _iter_batches(
    session: AsyncSession,
    since: datetime,
    until: datetime,
    blocked_only: bool,
    branch: str | None = None,
    client_ip: str | None = None,
    domain: str | None = None,
    domain_in: set[str] | None = None,
) -> AsyncIterator[Sequence[RawEvent]]:
    conditions = build_event_conditions(
        since,
        until,
        blocked_only=blocked_only,
        branch=branch,
        client_ip=client_ip,
        domain=domain,
        domain_in=domain_in,
    )
    last_id = 0
    while True:
        query = (
            select(RawEvent)
            .where(*conditions, RawEvent.id > last_id)
            .order_by(RawEvent.id)
            .limit(_STREAM_BATCH_SIZE)
        )
        rows = (await session.execute(query)).scalars().all()
        if not rows:
            return
        yield rows
        last_id = rows[-1].id
        if len(rows) < _STREAM_BATCH_SIZE:
            return


async def stream_csv(
    session: AsyncSession,
    since: datetime,
    until: datetime,
    blocked_only: bool,
    branch: str | None = None,
    row_counter: list[int] | None = None,
    client_ip: str | None = None,
    domain: str | None = None,
    domain_in: set[str] | None = None,
    columns: list[str] | None = None,
) -> AsyncIterator[str]:
    """row_counter, if given, is incremented (as a one-element out-param --
    an async generator can't return a value alongside its yields) by the
    number of rows in every batch actually written. export_job_service.run_job
    uses this instead of a separate COUNT(*) after the fact: since/until are
    a fixed range but this table keeps getting new matching rows inserted in
    real time, a follow-up count query can (and, under real traffic, will)
    see rows that arrived after streaming already finished, so it doesn't
    reliably describe what ended up in the file.

    `columns`, if given, must already be validated (see resolve_columns) --
    it's handed straight to csv.DictWriter as the column order. Extra keys
    on each row's dict (every column not selected) are silently dropped via
    extrasaction="ignore" rather than this function re-filtering every
    record itself."""
    resolved_columns = resolve_columns(columns)
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=resolved_columns, extrasaction="ignore")
    writer.writeheader()
    yield buffer.getvalue()

    batches = _iter_batches(session, since, until, blocked_only, branch, client_ip, domain, domain_in)
    async for batch in batches:
        if row_counter is not None:
            row_counter[0] += len(batch)
        buffer.seek(0)
        buffer.truncate(0)
        for row in batch:
            writer.writerow(_escape_record(_row_to_dict(row)))
        yield buffer.getvalue()


async def stream_json(
    session: AsyncSession,
    since: datetime,
    until: datetime,
    blocked_only: bool,
    branch: str | None = None,
    row_counter: list[int] | None = None,
    client_ip: str | None = None,
    domain: str | None = None,
    domain_in: set[str] | None = None,
    columns: list[str] | None = None,
) -> AsyncIterator[str]:
    """See stream_csv's docstring for row_counter and columns."""
    resolved_columns = resolve_columns(columns)
    yield "["
    first = True
    batches = _iter_batches(session, since, until, blocked_only, branch, client_ip, domain, domain_in)
    async for batch in batches:
        if row_counter is not None:
            row_counter[0] += len(batch)
        parts = []
        for row in batch:
            record = _row_to_dict(row)
            projected = {column: record[column] for column in resolved_columns}
            parts.append(("" if first else ",") + json.dumps(projected))
            first = False
        yield "".join(parts)
    yield "]"


async def download_csv(
    since: datetime,
    until: datetime,
    blocked_only: bool,
    branch: str | None = None,
    client_ip: str | None = None,
    domain: str | None = None,
    category: DomainCategoryLabel | None = None,
    columns: list[str] | None = None,
) -> AsyncIterator[str]:
    """Route-facing entry point for GET /api/export: opens and holds its own
    session for the whole streaming duration.

    A FastAPI `Depends(get_db)` session is the wrong tool here -- its
    cleanup runs right after the endpoint returns the StreamingResponse
    object, *before* Starlette actually pulls this generator to send the
    body, so a request-injected session would already be closed by the time
    any of the query above ran. Opening a fresh session inside the
    generator itself sidesteps that entirely.
    """
    async with db_module.AsyncSessionLocal() as session:
        domain_in = (
            await resolve_category_domains(session, since, until, category, branch)
            if category is not None
            else None
        )
        async for chunk in stream_csv(
            session, since, until, blocked_only, branch,
            client_ip=client_ip, domain=domain, domain_in=domain_in, columns=columns,
        ):
            yield chunk


async def download_json(
    since: datetime,
    until: datetime,
    blocked_only: bool,
    branch: str | None = None,
    client_ip: str | None = None,
    domain: str | None = None,
    category: DomainCategoryLabel | None = None,
    columns: list[str] | None = None,
) -> AsyncIterator[str]:
    """download_csv's JSON counterpart -- see its docstring for why this
    owns its own session instead of taking one via FastAPI's DI."""
    async with db_module.AsyncSessionLocal() as session:
        domain_in = (
            await resolve_category_domains(session, since, until, category, branch)
            if category is not None
            else None
        )
        async for chunk in stream_json(
            session, since, until, blocked_only, branch,
            client_ip=client_ip, domain=domain, domain_in=domain_in, columns=columns,
        ):
            yield chunk


# --- XLSX, job-only ---
#
# Unlike CSV/JSON (plain text, genuinely streamable one chunk at a time),
# .xlsx is a zip container whose central directory can only be written once
# every row is known -- openpyxl's write-only Workbook still keeps memory
# bounded (it spills worksheet data to a temp file as rows are appended,
# rather than holding every Cell object), but there's no way to hand
# Starlette a StreamingResponse that sends useful bytes before the very end.
# That's fine for the background-job path (already writes a complete file
# to disk before anything is served, see export_job_service.run_job) but
# wrong for GET /api/export's synchronous streaming contract, so xlsx is
# rejected there (see api/routes/export.py) rather than silently buffering
# an unbounded range in memory to fake the same interface.


def _sanitize_xlsx_value(value: object) -> object:
    """Excel's OOXML container rejects raw ASCII control characters (0x00-
    0x08, 0x0B-0x0C, 0x0E-0x1F) anywhere in a cell -- a Squid log line's url/
    user/etc. can contain one of these (a malformed request, or someone
    deliberately probing with control bytes), and unlike CSV/JSON, which
    don't care, openpyxl raises IllegalCharacterError and aborts the whole
    export the moment it hits one. Stripping them here (xlsx-only, not
    folded into _escape_record) means one odd byte in one row costs nothing
    more than that byte, instead of failing an export that may already be
    most of the way through a million rows."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


async def write_xlsx_rows(
    worksheet: Any,  # openpyxl.worksheet._write_only.WriteOnlyWorksheet -- see new_xlsx_workbook's own Any
    session: AsyncSession,
    since: datetime,
    until: datetime,
    blocked_only: bool,
    branch: str | None,
    client_ip: str | None,
    domain: str | None,
    domain_in: set[str] | None,
    columns: list[str],
    row_counter: list[int],
) -> AsyncIterator[None]:
    """Appends rows to an already-open write-only worksheet (header row
    already written by the caller) one DB batch at a time, yielding once per
    batch with no payload -- purely a checkpoint run_job hooks progress
    commits and cancellation checks onto, the same rhythm stream_csv/
    stream_json give it via their own per-chunk yields."""
    batches = _iter_batches(session, since, until, blocked_only, branch, client_ip, domain, domain_in)
    async for batch in batches:
        row_counter[0] += len(batch)
        for row in batch:
            record = _escape_record(_row_to_dict(row))
            worksheet.append([_sanitize_xlsx_value(record[column]) for column in columns])
        yield


# Wide enough that a typical value doesn't truncate/wrap by default --
# without this every column sits at Excel's ~8.43-character default, which
# for a `url`/`timestamp`-heavy export looks like everything ran together
# even though each value is already in its own cell.
_COLUMN_WIDTHS: dict[str, float] = {
    "id": 10,
    "timestamp": 30,
    "client_ip": 20,
    "branch": 12,
    "user": 14,
    "method": 10,
    "url": 50,
    "domain": 24,
    "action": 14,
    "status_code": 12,
    "bytes": 12,
    "blocked": 10,
}
_DEFAULT_COLUMN_WIDTH = 15


def new_xlsx_workbook(columns: list[str]) -> tuple[Workbook, Any]:
    """write_only=True is what keeps this bounded for a multi-million-row
    export -- openpyxl streams each appended row straight to a temp file
    instead of building the whole sheet as in-memory Cell objects, the xlsx
    equivalent of stream_csv's chunked StringIO. Column widths, a bold
    header, and freezing that header row are all still possible in this
    mode (unlike most per-cell styling, which write-only can't go back and
    apply after the fact) since they're set once, up front, rather than
    touching cells after they're written."""
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("events")
    for index, column in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = _COLUMN_WIDTHS.get(
            column, _DEFAULT_COLUMN_WIDTH
        )
    worksheet.freeze_panes = "A2"
    header_cells = [WriteOnlyCell(worksheet, value=column) for column in columns]
    for cell in header_cells:
        cell.font = Font(bold=True)
    worksheet.append(header_cells)
    return workbook, worksheet
