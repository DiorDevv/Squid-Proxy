import csv
import io
import json
from datetime import UTC, datetime

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain_category import DomainCategoryLabel
from app.models.raw_event import RawEvent
from app.schemas.common import RangeParam
from app.services.export_service import (
    EXPORT_COLUMNS,
    EXPORT_ROW_LIMIT,
    download_csv,
    export_as_csv,
    export_as_json,
    new_xlsx_workbook,
    resolve_columns,
    stream_csv,
    stream_json,
    write_xlsx_rows,
)


def _make_event(**overrides) -> RawEvent:
    defaults = dict(
        timestamp=datetime.now(UTC),
        duration_ms=1,
        client_ip="10.0.0.1",
        action="TCP_MISS",
        status_code=200,
        bytes=100,
        method="GET",
        url="http://example.com/",
        domain="example.com",
        user="alice",
        hierarchy="HIER_DIRECT",
        peer=None,
        content_type="text/html",
        blocked=False,
    )
    defaults.update(overrides)
    return RawEvent(**defaults)


async def test_export_as_csv_escapes_formula_injection(db_session: AsyncSession):
    db_session.add_all(
        [
            _make_event(client_ip="10.0.0.1", domain="=cmd|'/c calc'!A1", url="+HYPERLINK(\"http://evil\")"),
            _make_event(client_ip="10.0.0.2", user="@SUM(1+1)"),
        ]
    )
    await db_session.commit()

    csv_body = await export_as_csv(db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False)
    reader = csv.DictReader(io.StringIO(csv_body))
    rows_by_ip = {row["client_ip"]: row for row in reader}

    assert rows_by_ip["10.0.0.1"]["domain"].startswith("'=")
    assert rows_by_ip["10.0.0.1"]["url"].startswith("'+")
    assert rows_by_ip["10.0.0.2"]["user"].startswith("'@")
    # Safe values are left untouched.
    assert rows_by_ip["10.0.0.2"]["client_ip"] == "10.0.0.2"


async def test_export_as_json_does_not_mangle_values(db_session: AsyncSession):
    db_session.add(_make_event(domain="=evil.com"))
    await db_session.commit()

    body = await export_as_json(db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False)
    parsed = json.loads(body)

    # JSON isn't opened in a spreadsheet, so it shouldn't be quote-prefixed.
    assert parsed[0]["domain"] == "=evil.com"


async def test_export_respects_blocked_only_filter(db_session: AsyncSession):
    db_session.add_all(
        [
            _make_event(client_ip="10.0.0.1", blocked=False),
            _make_event(client_ip="10.0.0.2", blocked=True, action="TCP_DENIED", status_code=403),
        ]
    )
    await db_session.commit()

    csv_body = await export_as_csv(db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=True)
    reader = csv.DictReader(io.StringIO(csv_body))
    rows = list(reader)

    assert len(rows) == 1
    assert rows[0]["client_ip"] == "10.0.0.2"


async def test_export_row_count_is_capped(db_session: AsyncSession, monkeypatch):
    import app.services.export_service as export_service_module

    monkeypatch.setattr(export_service_module, "EXPORT_ROW_LIMIT", 3)
    db_session.add_all([_make_event(client_ip=f"10.0.0.{i}") for i in range(10)])
    await db_session.commit()

    csv_body = await export_as_csv(db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False)
    reader = csv.DictReader(io.StringIO(csv_body))
    assert len(list(reader)) == 3
    # The real constant is untouched by the monkeypatch above.
    assert EXPORT_ROW_LIMIT == 100_000


async def test_stream_csv_is_not_row_limited(db_session: AsyncSession, monkeypatch):
    # Force many small batches (instead of one query returning everything)
    # to prove pagination across batch boundaries doesn't drop or duplicate
    # rows, and that -- unlike export_as_csv -- this never consults
    # EXPORT_ROW_LIMIT at all. GET /api/export and
    # scripts/archive_weekly_export.py both rely on this for ranges far
    # past that cap (a real deployment's raw_events table holds millions of
    # rows for even a single day).
    import app.services.export_service as export_service_module

    monkeypatch.setattr(export_service_module, "_STREAM_BATCH_SIZE", 3)
    row_total = 250
    db_session.add_all([_make_event(client_ip=f"10.0.{i // 250}.{i % 250}") for i in range(row_total)])
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in stream_csv(db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False)
    ]
    csv_body = "".join(chunks)
    reader = csv.DictReader(io.StringIO(csv_body))
    assert len(list(reader)) == row_total


async def test_stream_csv_escapes_formula_injection(db_session: AsyncSession):
    db_session.add(_make_event(client_ip="10.0.0.1", domain="=cmd|'/c calc'!A1"))
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in stream_csv(db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False)
    ]
    reader = csv.DictReader(io.StringIO("".join(chunks)))
    assert next(reader)["domain"].startswith("'=")


async def test_stream_json_produces_valid_json_across_batches(db_session: AsyncSession, monkeypatch):
    import app.services.export_service as export_service_module

    monkeypatch.setattr(export_service_module, "_STREAM_BATCH_SIZE", 2)
    db_session.add_all([_make_event(client_ip=f"10.0.0.{i}") for i in range(7)])
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in stream_json(
            db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False
        )
    ]
    parsed = json.loads("".join(chunks))
    assert len(parsed) == 7


async def test_export_route_returns_csv_with_headers(app_client: AsyncClient, admin_token, auth_headers):
    response = await app_client.get("/api/export?format=csv", headers=auth_headers(admin_token))
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "attachment" in response.headers["content-disposition"]


async def test_export_route_returns_json(app_client: AsyncClient, admin_token, auth_headers):
    response = await app_client.get("/api/export?format=json", headers=auth_headers(admin_token))
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/json")
    assert json.loads(response.text) == []


async def test_export_route_rejects_xlsx(app_client: AsyncClient, admin_token, auth_headers):
    # xlsx can't be produced as a genuine incremental stream (see
    # export_service.py's "XLSX, job-only" section) -- only reachable via
    # the background job path.
    response = await app_client.get("/api/export?format=xlsx", headers=auth_headers(admin_token))
    assert response.status_code == 400
    assert "export/jobs" in response.json()["detail"]


def test_resolve_columns_returns_all_by_default():
    assert resolve_columns(None) == EXPORT_COLUMNS


def test_resolve_columns_preserves_requested_order():
    assert resolve_columns(["domain", "id"]) == ["domain", "id"]


def test_resolve_columns_rejects_unknown_column():
    with pytest.raises(ValueError, match="not-a-real-column"):
        resolve_columns(["id", "not-a-real-column"])


def test_resolve_columns_rejects_empty_list():
    with pytest.raises(ValueError):
        resolve_columns([])


async def test_stream_csv_projects_selected_columns_only(db_session: AsyncSession):
    db_session.add(_make_event(client_ip="10.0.0.1", domain="example.com"))
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in stream_csv(
            db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False,
            columns=["client_ip", "domain"],
        )
    ]
    reader = csv.DictReader(io.StringIO("".join(chunks)))
    row = next(reader)
    assert reader.fieldnames == ["client_ip", "domain"]
    assert row == {"client_ip": "10.0.0.1", "domain": "example.com"}


async def test_stream_json_projects_selected_columns_only(db_session: AsyncSession):
    db_session.add(_make_event(client_ip="10.0.0.1", domain="example.com"))
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in stream_json(
            db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False,
            columns=["client_ip", "domain"],
        )
    ]
    parsed = json.loads("".join(chunks))
    assert parsed == [{"client_ip": "10.0.0.1", "domain": "example.com"}]


async def test_stream_csv_respects_client_ip_filter(db_session: AsyncSession):
    db_session.add_all(
        [
            _make_event(client_ip="10.0.0.1"),
            _make_event(client_ip="10.0.0.2"),
        ]
    )
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in stream_csv(
            db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False,
            client_ip="10.0.0.2",
        )
    ]
    reader = csv.DictReader(io.StringIO("".join(chunks)))
    rows = list(reader)
    assert len(rows) == 1
    assert rows[0]["client_ip"] == "10.0.0.2"


async def test_download_csv_respects_category_filter(db_session: AsyncSession, monkeypatch):
    # example.com isn't in the built-in curated category list, so it falls
    # back to UNCATEGORIZED via category_inference.infer_category --
    # exercising resolve_category_domains without needing a DomainCategory
    # override row for either domain.
    import app.models.db as db_module

    monkeypatch.setattr(db_module, "AsyncSessionLocal", lambda: db_session)

    db_session.add_all(
        [
            _make_event(client_ip="10.0.0.1", domain="example.com"),
            _make_event(client_ip="10.0.0.2", domain="facebook.com"),
        ]
    )
    await db_session.commit()

    chunks = [
        chunk
        async for chunk in download_csv(
            RangeParam.ONE_HOUR.since(), datetime.now(UTC), blocked_only=False,
            category=DomainCategoryLabel.SOCIAL_MEDIA,
        )
    ]

    reader = csv.DictReader(io.StringIO("".join(chunks)))
    rows = list(reader)
    assert len(rows) == 1
    assert rows[0]["domain"] == "facebook.com"


async def test_write_xlsx_rows_produces_readable_workbook(db_session: AsyncSession, tmp_path):
    db_session.add_all(
        [
            _make_event(client_ip="10.0.0.1", domain="example.com"),
            _make_event(client_ip="10.0.0.2", domain="=evil.com"),  # formula-injection risk
        ]
    )
    await db_session.commit()

    columns = ["client_ip", "domain"]
    workbook, worksheet = new_xlsx_workbook(columns)
    row_counter = [0]
    async for _ in write_xlsx_rows(
        worksheet, db_session, RangeParam.ONE_HOUR.since(), datetime.now(UTC), False,
        None, None, None, None, columns, row_counter,
    ):
        pass
    out_path = tmp_path / "events.xlsx"
    workbook.save(str(out_path))
    workbook.close()

    assert row_counter[0] == 2
    loaded = load_workbook(out_path, read_only=True)
    ws = loaded["events"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("client_ip", "domain")
    assert ("10.0.0.1", "example.com") in rows
    # Formula-injection escaping applies to xlsx cells too, not just CSV.
    assert ("10.0.0.2", "'=evil.com") in rows
